import requests
import openpyxl

def get_last_new_rating(username):
    url = f"https://codeforces.com/api/user.rating?handle={username.strip()}"
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        if 'result' in data:
            ratings = data['result']
            if ratings:
                return ratings[-1]['newRating']
    return None

def main():
    workbook = openpyxl.load_workbook('Book1.xlsx')
    sheet = workbook.active
    max_row = sheet.max_row

    for row_num in range(1, max_row + 1):
        username = sheet.cell(row=row_num, column=1).value
        if username:
            rating = get_last_new_rating(username)
            if rating is not None:
                sheet.cell(row=row_num, column=2).value = rating
            else:
                print(f"Failed to fetch rating for {username}")

    workbook.save('usernames_with_ratings.xlsx')

if __name__ == "__main__":
    main()
